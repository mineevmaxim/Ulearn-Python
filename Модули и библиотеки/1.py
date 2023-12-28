import csv
from openpyxl import Workbook


def create_report():
    with open('vacancies.csv', 'r', encoding='utf-8-sig') as csvfile:
        reader = csv.reader(csvfile)

        data_by_year = {}
        data_by_city = {}

        for row in reader:
            year = int(row[5].split('-')[0])
            city = row[4]
            salary = (float(row[1]) + float(row[2])) / 2

            if year in data_by_year:
                data_by_year[year]['salary'].append(salary)
                data_by_year[year]['vacancy'] += 1
            else:
                data_by_year[year] = {'salary': [salary], 'vacancy': 1}

            if city in data_by_city:
                data_by_city[city]['salary'].append(salary)
                data_by_city[city]['vacancy'] += 1
            else:
                data_by_city[city] = {'salary': [salary], 'vacancy': 1}

    workbook = Workbook()
    year_sheet = workbook.create_sheet(title="Статистика по годам")

    year_sheet['A1'] = "Год"
    year_sheet['B1'] = "Средняя зарплата"
    year_sheet['C1'] = "Количество вакансий"

    sorted_data_by_year = sorted_data_by_city = sorted(data_by_year.items(), key=lambda x: x[0])

    row_num = 2
    for year, data in sorted_data_by_year:
        year_sheet.cell(row=row_num, column=1, value=year)
        year_sheet.cell(row=row_num, column=2, value=round(sum(data['salary']) / len(data['salary'])))
        year_sheet.cell(row=row_num, column=3, value=data['vacancy'])
        row_num += 1

    city_sheet = workbook.create_sheet(title="Статистика по городам")

    city_sheet['A1'] = "Город"
    city_sheet['B1'] = "Уровень зарплат"
    city_sheet['D1'] = "Город"
    city_sheet['E1'] = "Доля вакансий, %"

    filtered_data_by_city = {city: data for city, data in data_by_city.items() if
                             data['vacancy'] / sum([item['vacancy'] for item in data_by_city.values()]) > 0.01}

    sorted_data_by_city = sorted(filtered_data_by_city.items(),
                                 key=lambda x: (sum(x[1]['salary']) / len(x[1]['salary']), x[1]['vacancy']),
                                 reverse=True)

    row_num = 2
    for i in range(min(len(sorted_data_by_city), 10)):
        city = sorted_data_by_city[i][0]
        data = sorted_data_by_city[i][1]
        city_sheet.cell(row=row_num, column=1, value=city)
        city_sheet.cell(row=row_num, column=2, value=round(sum(data['salary']) / len(data['salary'])))
        row_num += 1

    vacancies_count = sum([item['vacancy'] for item in data_by_city.values()])
    sorted_data_by_city = sorted(filtered_data_by_city.items(),
                                 key=lambda x: (1 / (int(x[1]['vacancy']) / vacancies_count), x[0]), reverse=False)

    row_num = 2
    for i in range(min(len(sorted_data_by_city), 10)):
        city = sorted_data_by_city[i][0]
        data = sorted_data_by_city[i][1]
        city_sheet.cell(row=row_num, column=4, value=city)
        city_sheet.cell(row=row_num, column=5,
                        value=round(data['vacancy'] / sum([item['vacancy'] for item in data_by_city.values()]) * 100,
                                    2))
        row_num += 1

    del workbook["Sheet"]

    workbook.save("student_works/report.xlsx")


create_report()